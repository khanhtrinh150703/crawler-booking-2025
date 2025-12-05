# core/excel_exporter.py (đã cập nhật)

import os
import pandas as pd
from error.check_json import process_json_file
from config.config import ROOT_DIR, EXCEL_DIR, EXCEL_NAME
from utils.helpers import ensure_dir


def collect_excel_data(path=ROOT_DIR, prefix="", data=None):
    if data is None:
        data = []
    for item in sorted(os.listdir(path)):
        item_path = os.path.join(path, item)
        if os.path.isdir(item_path):
            # Xác định tên tỉnh/thành phố từ tên thư mục (mức 1 ngay dưới ROOT_DIR)
            new_prefix = f"{prefix}{item}/" if prefix else f"{item}/"
            collect_excel_data(item_path, new_prefix, data)
        elif item.endswith(".json"):
            has_error, reason, review_count, hotel_name, file_name, total_rating, viet_positive_count = process_json_file(item_path)
            
            rel_path = f"{prefix}{file_name}"
            status = "LỖI" if has_error else "HỢP LỆ"
            data.append({
                "File Path": rel_path,
                "Tên File": file_name,
                "Tên Khách Sạn": hotel_name,
                "Tổng Reviews Cào Được": review_count,
                "Total Rating": total_rating if total_rating is not None else "",
                "VN có comment+": viet_positive_count,
                "Trạng Thái": status,
                "Lỗi (nếu có)": reason or "",
                # Thêm cột ẩn để dễ group theo tỉnh sau này
                "_Tỉnh": prefix.split("/")[0] if prefix else "Unknown",
            })
    return data

def export_to_excel():
    ensure_dir(EXCEL_DIR)
    print("\nĐang thu thập dữ liệu để xuất Excel...")
    raw_data = collect_excel_data()
    if not raw_data:
        print("Không có dữ liệu để xuất Excel.")
        return

    df = pd.DataFrame(raw_data)
    
    # ===== Sheet 1: Báo cáo chi tiết =====
    df_report = df.drop(columns=["_Tỉnh"])

    # ===== Sheet 2: Thống kê theo Tỉnh + Dòng Tổng cộng =====
    df_province = df.groupby("_Tỉnh", as_index=False).agg(
        Tổng_Reviews_Cào_Được=("Tổng Reviews Cào Được", "sum"),
        Review_Tiếng_Việt_Comment_Positive=("VN có comment+", "sum"),
        Số_Khách_Sạn=("Tên Khách Sạn", "count")
    ).rename(columns={"_Tỉnh": "Tỉnh/Thành phố"})
    
    df_province = df_province.sort_values("Tổng_Reviews_Cào_Được", ascending=False)

    # Dòng tổng cộng
    total_row = pd.DataFrame([{
        "Tỉnh/Thành phố": "TỔNG CỘNG",
        "Tổng_Reviews_Cào_Được": df_province["Tổng_Reviews_Cào_Được"].sum(),
        "Review_Tiếng_Việt_Comment_Positive": df_province["Review_Tiếng_Việt_Comment_Positive"].sum(),
        "Số_Khách_Sạn": df_province["Số_Khách_Sạn"].sum(),
    }])

    df_province_final = pd.concat([df_province, total_row], ignore_index=True)

    excel_path = os.path.join(EXCEL_DIR, EXCEL_NAME)

    try:
        # === MỞ ExcelWriter một lần duy nhất ===
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            
            # ── Sheet 1: Hotels Report ──
            df_report.to_excel(writer, sheet_name='Hotels Report', index=False)
            ws1 = writer.sheets['Hotels Report']
            for i, col in enumerate(df_report.columns):
                max_len = max(df_report[col].astype(str).map(len).max(), len(col)) + 2
                ws1.column_dimensions[chr(65 + i)].width = min(max_len, 50)

            # ── Sheet 2: Thống kê theo Tỉnh (có Tổng cộng) ──
            df_province_final.to_excel(writer, sheet_name='Thống kê theo Tỉnh', index=False)
            ws2 = writer.sheets['Thống kê theo Tỉnh']
            
            # Điều chỉnh độ rộng cột
            for i, col in enumerate(df_province_final.columns):
                max_len = max(df_province_final[col].astype(str).map(len).max(), len(col)) + 2
                ws2.column_dimensions[chr(65 + i)].width = min(max_len, 50)

            # Định dạng số
            from openpyxl.styles import Font, PatternFill
            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=2, max_col=4):
                for cell in row:
                    cell.number_format = '#,##0'

            # Tô màu + in đậm dòng TỔNG CỘNG
            total_row_idx = ws2.max_row
            for cell in ws2[total_row_idx]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")

        # ── In thông báo thành công ──
        total_vn_pos = df["VN có comment+"].sum()
        total_reviews = df["Tổng Reviews Cào Được"].sum()
        print(f"Xuất Excel thành công: {excel_path}")
        print(f"   → Tổng cộng: {len(raw_data)} khách sạn")
        print(f"   → Tổng review cào được: {total_reviews:,}")
        print(f"   → Review tiếng Việt có comment+: {total_vn_pos:,}")
        print(f"   → Có 2 sheet, trong đó 'Thống kê theo Tỉnh' đã có dòng TỔNG CỘNG")

    except ImportError:
        print("Cảnh báo: Thiếu 'pandas' hoặc 'openpyxl'. Cài đặt bằng: pip install pandas openpyxl")
    except Exception as e:
        print(f"Lỗi khi xuất Excel: {e}")